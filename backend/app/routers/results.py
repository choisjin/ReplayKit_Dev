"""Test results API routes."""

import html as _html
import io
import json
import os
import shutil
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse

router = APIRouter(prefix="/api/results", tags=["results"])

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent  # server.py 위치
RESULTS_DIR = Path(__file__).resolve().parent.parent.parent / "results"
SCREENSHOTS_DIR = Path(__file__).resolve().parent.parent.parent / "screenshots"
RECORDINGS_DIR = _PROJECT_ROOT / "Results" / "Video"
EXPORT_ROOT = _PROJECT_ROOT / "Results"
_TOOLS_DIR = _PROJECT_ROOT / "tools"


def _find_ffmpeg() -> str | None:
    """ffmpeg 실행 파일 경로를 반환. 시스템 PATH → tools/ 폴더 순으로 탐색."""
    # 시스템 PATH
    found = shutil.which("ffmpeg")
    if found:
        return found
    # 프로젝트 tools/ 폴더
    local = _TOOLS_DIR / "ffmpeg.exe"
    if local.is_file():
        return str(local)
    # tools/ffmpeg/bin/ 구조 (일반적인 ffmpeg 배포 패키지)
    local_bin = _TOOLS_DIR / "ffmpeg" / "bin" / "ffmpeg.exe"
    if local_bin.is_file():
        return str(local_bin)
    return None


@router.get("/list")
async def list_results():
    """List all test result files (런 폴더 + 레거시 플랫 파일 모두 탐색)."""
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    results = []
    seen: set[str] = set()

    # 1) 런 폴더: results/{ts}_{scenario}/result.json
    for d in sorted(RESULTS_DIR.iterdir(), reverse=True):
        if not d.is_dir():
            continue
        rj = d / "result.json"
        if not rj.exists():
            continue
        try:
            data = json.loads(rj.read_text(encoding="utf-8"))
        except Exception:
            continue
        key = d.name
        seen.add(key)
        results.append({
            "filename": f"{d.name}/result.json",
            "run_folder": d.name,
            "scenario_name": data.get("scenario_name", ""),
            "status": data.get("status", ""),
            "total_steps": data.get("total_steps", 0),
            "total_repeat": data.get("total_repeat", 1),
            "passed_steps": data.get("passed_steps", 0),
            "failed_steps": data.get("failed_steps", 0),
            "warning_steps": data.get("warning_steps", 0),
            "error_steps": data.get("error_steps", 0),
            "started_at": data.get("started_at", ""),
            "finished_at": data.get("finished_at", ""),
        })

    # 2) 레거시 플랫: results/*.json
    for f in sorted(RESULTS_DIR.glob("*.json"), reverse=True):
        data = json.loads(f.read_text(encoding="utf-8"))
        results.append({
            "filename": f.name,
            "run_folder": "",
            "scenario_name": data.get("scenario_name", ""),
            "status": data.get("status", ""),
            "total_steps": data.get("total_steps", 0),
            "total_repeat": data.get("total_repeat", 1),
            "passed_steps": data.get("passed_steps", 0),
            "failed_steps": data.get("failed_steps", 0),
            "warning_steps": data.get("warning_steps", 0),
            "error_steps": data.get("error_steps", 0),
            "started_at": data.get("started_at", ""),
            "finished_at": data.get("finished_at", ""),
        })
    return {"results": results}


def _resolve_image_path(rel_path: str | None) -> Path | None:
    """Resolve a relative screenshot path to an absolute filesystem path."""
    if not rel_path:
        return None
    # Handle absolute paths from older results
    p = rel_path.replace("\\", "/")
    idx = p.find("/screenshots/")
    if idx >= 0:
        p = p[idx + len("/screenshots/"):]
    full = SCREENSHOTS_DIR / p
    return full if full.exists() else None


def _html_image_src(rel_path: str | None, html_dir: Path) -> str:
    """저장된 이미지 경로(RESULTS_DIR 또는 SCREENSHOTS_DIR 기준)를 HTML이 위치한
    디렉토리 기준 상대 경로로 변환. 파일이 존재하지 않으면 빈 문자열."""
    if not rel_path:
        return ""
    rp = str(rel_path).replace("\\", "/")
    # /screenshots/ 프리픽스가 포함된 경우 제거
    idx = rp.find("/screenshots/")
    if idx >= 0:
        rp_ss = rp[idx + len("/screenshots/"):]
    else:
        rp_ss = rp
    candidates = [
        RESULTS_DIR / rp,
        SCREENSHOTS_DIR / rp_ss,
    ]
    for abs_path in candidates:
        try:
            if abs_path.exists():
                try:
                    return abs_path.relative_to(html_dir).as_posix()
                except ValueError:
                    return os.path.relpath(str(abs_path), str(html_dir)).replace("\\", "/")
        except OSError:
            continue
    return ""


_HTML_STYLE = """
*,*::before,*::after { box-sizing: border-box; }
body { font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  margin: 0; padding: 16px 20px; font-size: 13px; color: #1a1a2e; background: #f5f7fa; }
h1 { font-size: 20px; font-weight: 700; margin: 0 0 6px; color: #1a1a2e; letter-spacing: -0.3px; }
.meta { display: flex; flex-wrap: wrap; gap: 6px 16px; margin: 0 0 14px; color: #555; font-size: 12px; }
.meta .badge { display: inline-flex; align-items: center; gap: 4px; padding: 2px 10px;
  border-radius: 12px; font-weight: 600; font-size: 12px; }
.meta .badge.pass { background: #dcfce7; color: #15803d; }
.meta .badge.fail { background: #fee2e2; color: #b91c1c; }
.meta .badge.error { background: #ffedd5; color: #9a3412; }
.meta .stat { padding: 2px 8px; border-radius: 4px; background: #e8ecf1; font-weight: 500; }
.controls { position: sticky; top: 0; z-index: 20; background: #fff; padding: 10px 16px;
  border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); margin-bottom: 12px;
  display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
.controls input[type="text"] { font-size: 12px; padding: 6px 10px; border: 1px solid #d1d5db;
  border-radius: 6px; min-width: 260px; outline: none; transition: border-color 0.15s; }
.controls input[type="text"]:focus { border-color: #3b82f6; box-shadow: 0 0 0 2px rgba(59,130,246,0.15); }
.controls button { font-size: 12px; padding: 6px 14px; border: none; border-radius: 6px;
  cursor: pointer; font-weight: 500; transition: all 0.15s; }
.controls .btn-primary { background: #3b82f6; color: #fff; }
.controls .btn-primary:hover { background: #2563eb; }
.controls .btn-secondary { background: #f3f4f6; color: #374151; border: 1px solid #d1d5db; }
.controls .btn-secondary:hover { background: #e5e7eb; }
.controls .count { margin-left: auto; color: #6b7280; font-size: 12px; }
/* Tabulator 테마 오버라이드 */
.tabulator { border: none; border-radius: 8px; overflow: hidden;
  box-shadow: 0 1px 4px rgba(0,0,0,0.08); font-size: 12px; background: #fff; }
.tabulator .tabulator-header { background: #1e3a5f; color: #fff; font-weight: 600;
  border-bottom: 2px solid #15294a; }
.tabulator .tabulator-header .tabulator-col { background: transparent; color: #fff;
  border-right: 1px solid rgba(255,255,255,0.1); }
.tabulator .tabulator-header .tabulator-col.tabulator-sortable:hover { background: rgba(255,255,255,0.1); }
.tabulator .tabulator-header .tabulator-col .tabulator-col-title { color: #fff; padding: 8px 6px;
  font-size: 11px; text-transform: uppercase; letter-spacing: 0.3px; }
.tabulator .tabulator-header .tabulator-col .tabulator-col-sorter { color: rgba(255,255,255,0.5); }
.tabulator .tabulator-header .tabulator-header-filter { padding: 4px 4px 6px; }
.tabulator .tabulator-header .tabulator-header-filter input,
.tabulator .tabulator-header .tabulator-header-filter select {
  font-size: 11px; padding: 4px 6px; border: 1px solid #d1d5db; border-radius: 4px;
  background: #fff; color: #374151; width: 100%; outline: none; }
.tabulator .tabulator-header .tabulator-header-filter input:focus,
.tabulator .tabulator-header .tabulator-header-filter select:focus {
  border-color: #60a5fa; box-shadow: 0 0 0 2px rgba(96,165,250,0.2); }
.tabulator .tabulator-tableholder { background: #fff; }
.tabulator-row { border-bottom: 1px solid #f0f0f0; }
.tabulator-row .tabulator-cell { padding: 4px 6px; border-right: 1px solid #f3f4f6; }
.tabulator-row.tabulator-row-even { background: #fafbfc; }
.tabulator-row:hover { background: #eff6ff !important; }
.tabulator-row.tabulator-row-even:hover { background: #eff6ff !important; }
/* Status 배지 */
.st-badge { display: inline-block; padding: 3px 10px; border-radius: 10px; font-size: 11px;
  font-weight: 700; letter-spacing: 0.3px; text-transform: uppercase; }
.st-badge.pass { background: #dcfce7; color: #15803d; }
.st-badge.fail { background: #fee2e2; color: #b91c1c; }
.st-badge.warning { background: #fef9c3; color: #854d0e; }
.st-badge.error { background: #ffedd5; color: #9a3412; }
.img-thumb { max-width: 170px; max-height: 130px; display: block; margin: 0 auto;
  cursor: pointer; border-radius: 4px; border: 1px solid #e5e7eb; transition: transform 0.15s; }
.img-thumb:hover { transform: scale(1.03); box-shadow: 0 2px 8px rgba(0,0,0,0.12); }
/* 이미지 프리뷰 */
.preview-overlay { position: fixed; inset: 0; background: rgba(0,0,0,0.88); display: none;
  align-items: center; justify-content: center; z-index: 9999; cursor: zoom-out; }
.preview-overlay.open { display: flex; }
.preview-overlay img { max-width: 95vw; max-height: 95vh; border-radius: 6px;
  box-shadow: 0 8px 40px rgba(0,0,0,0.4); }
@media print {
  body { margin: 8px; padding: 0; background: #fff; }
  .controls { display: none !important; }
  .tabulator .tabulator-header .tabulator-header-filter { display: none !important; }
  .tabulator { box-shadow: none; border: 1px solid #ccc; height: auto !important;
    max-height: none !important; overflow: visible !important; }
  .tabulator .tabulator-tableholder { height: auto !important; max-height: none !important;
    overflow: visible !important; }
  .tabulator-row { page-break-inside: avoid; }
  .img-thumb { max-width: 140px; max-height: 110px; border: none; }
}
"""

# Tabulator 기반 리포트 초기화 스크립트 — 데이터는 window.__REPORT_DATA__ 전역에서 읽는다.
_HTML_SCRIPT = r"""
(function(){
  /* ---------- 셀 포맷터 ---------- */
  function statusFmt(cell){
    var v = (cell.getValue() || '').toString().toLowerCase();
    return '<span class="st-badge ' + v + '">' + v.toUpperCase() + '</span>';
  }
  function imgFmt(cell){
    var v = cell.getValue();
    if (!v) return '<span style="color:#bbb">—</span>';
    return '<img class="img-thumb" loading="lazy" src="' + v + '" alt="">';
  }

  /* ---------- 고유값 수집 헬퍼 ---------- */
  function uniqueVals(data, field){
    var seen = {};
    var out = [];
    for (var i = 0; i < data.length; i++){
      var v = data[i][field];
      if (v == null || v === '') continue;
      var s = String(v);
      if (!seen[s]) { seen[s] = true; out.push(s); }
    }
    out.sort(function(a,b){ return a.localeCompare(b, undefined, {numeric:true}); });
    return out;
  }

  /* ---------- 드롭다운 헤더 필터 에디터 (리스트형) ---------- */
  function listEditor(field, placeholder){
    return function(cell, onRendered, success, cancel, editorParams){
      var vals = editorParams.values || [];
      var select = document.createElement('select');
      select.style.cssText = 'width:100%; font-size:11px; padding:3px 4px; border:1px solid #d1d5db; border-radius:4px; background:#fff; color:#374151; cursor:pointer;';
      var opt0 = document.createElement('option');
      opt0.value = ''; opt0.textContent = placeholder || '전체';
      select.appendChild(opt0);
      for (var i = 0; i < vals.length; i++){
        var opt = document.createElement('option');
        opt.value = vals[i]; opt.textContent = vals[i];
        select.appendChild(opt);
      }
      var cur = cell.getValue();
      if (cur) select.value = cur;
      select.addEventListener('change', function(){ success(select.value); });
      return select;
    };
  }

  /* ---------- 이미지 프리뷰 ---------- */
  function onImgClick(e){
    if (e.target && e.target.classList.contains('img-thumb')){
      document.getElementById('preview-img').src = e.target.src;
      document.getElementById('preview-overlay').classList.add('open');
    }
  }
  function closePreview(){
    document.getElementById('preview-overlay').classList.remove('open');
    document.getElementById('preview-img').src = '';
  }

  /* ---------- 컬럼 정의 ---------- */
  function buildColumns(data){
    var tsVals   = uniqueVals(data, 'timestamp');
    var cyVals   = uniqueVals(data, 'cycle');
    var stepVals = uniqueVals(data, 'step_id');
    var devVals  = uniqueVals(data, 'device');
    var delVals  = uniqueVals(data, 'delay');
    var durVals  = uniqueVals(data, 'duration');

    var va = "middle";
    return [
      { title:"Time Stamp", field:"timestamp", width:150, vertAlign:va,
        headerFilter:listEditor('timestamp','전체'), headerFilterParams:{values:tsVals}, headerFilterFunc:"=" },
      { title:"Cycle", field:"cycle", width:70, hozAlign:"center", vertAlign:va,
        headerFilter:listEditor('cycle','전체'), headerFilterParams:{values:cyVals}, headerFilterFunc:"=" },
      { title:"Step", field:"step_id", width:70, hozAlign:"center", vertAlign:va,
        headerFilter:listEditor('step_id','전체'), headerFilterParams:{values:stepVals}, headerFilterFunc:"=" },
      { title:"Device", field:"device", width:120, hozAlign:"center", vertAlign:va,
        headerFilter:listEditor('device','전체'), headerFilterParams:{values:devVals}, headerFilterFunc:"=" },
      { title:"Command", field:"command", widthGrow:2, vertAlign:va,
        headerFilter:listEditor('command','전체'), headerFilterParams:{values:uniqueVals(data,'command')}, headerFilterFunc:"=" },
      { title:"Remark", field:"description", widthGrow:2, vertAlign:va,
        headerFilter:listEditor('description','전체'), headerFilterParams:{values:uniqueVals(data,'description')}, headerFilterFunc:"=" },
      { title:"Status", field:"status", width:90, hozAlign:"center", vertAlign:va, formatter:statusFmt,
        headerFilter:listEditor('status','전체'), headerFilterParams:{values:["pass","fail","warning","error"]}, headerFilterFunc:"=" },
      { title:"Delay", field:"delay", width:80, hozAlign:"center", vertAlign:va,
        headerFilter:listEditor('delay','전체'), headerFilterParams:{values:delVals}, headerFilterFunc:"=" },
      { title:"Duration", field:"duration", width:90, hozAlign:"center", vertAlign:va,
        headerFilter:listEditor('duration','전체'), headerFilterParams:{values:durVals}, headerFilterFunc:"=" },
      { title:"Expected", field:"expected_src", width:200, hozAlign:"center", vertAlign:va,
        formatter:imgFmt, headerSort:false },
      { title:"Actual", field:"actual_src", width:200, hozAlign:"center", vertAlign:va,
        formatter:imgFmt, headerSort:false },
    ];
  }

  /* ---------- 전역 검색 필터 ---------- */
  var TEXT_FIELDS = ["timestamp","cycle","step_id","device","command","description","status","delay","duration"];
  function globalFilter(row, params){
    var q = params.q;
    if (!q) return true;
    for (var i = 0; i < TEXT_FIELDS.length; i++){
      var v = row[TEXT_FIELDS[i]];
      if (v != null && String(v).toLowerCase().indexOf(q) !== -1) return true;
    }
    return false;
  }

  /* ---------- 초기화 ---------- */
  function init(){
    var data = (window.__REPORT_DATA__ && window.__REPORT_DATA__.rows) || [];
    document.getElementById('filter-total').textContent = data.length;

    var table = new Tabulator("#results-table", {
      data: data,
      columns: buildColumns(data),
      layout: "fitDataStretch",
      maxHeight: "calc(100vh - 170px)",
      renderVertical: "basic",
      placeholder: "표시할 결과가 없습니다",
      headerSortClickElement: "icon",
      rowHeight: false,
      cellVertAlign: "middle",
    });
    window.__table = table;

    table.on("dataFiltered", function(filters, rows){
      document.getElementById('filter-visible').textContent = rows.length;
    });

    // 전역 검색
    var gEl = document.getElementById('filter-text');
    gEl.addEventListener('input', function(){
      var v = (gEl.value || '').trim().toLowerCase();
      table.removeFilter(globalFilter);
      if (v) table.addFilter(globalFilter, { q: v });
    });

    // 필터 초기화
    document.getElementById('filter-reset').addEventListener('click', function(){
      table.clearHeaderFilter();
      gEl.value = '';
      table.removeFilter(globalFilter);
    });

    // PDF 저장 — window.print()로 전체 행 출력 (가상 렌더링 우회)
    document.getElementById('pdf-btn').addEventListener('click', function(){ window.print(); });

    // 이미지 프리뷰
    document.getElementById('results-table').addEventListener('click', onImgClick);
    document.getElementById('preview-overlay').addEventListener('click', closePreview);
    document.addEventListener('keydown', function(e){ if (e.key === 'Escape') closePreview(); });
  }

  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', init);
  else init();
})();
"""


def _build_html_report(data: dict, output_path: Path) -> str:
    """Tabulator 기반 경량 HTML 리포트.

    데이터는 window.__REPORT_DATA__에 JSON으로 임베드되고,
    Tabulator가 열별 필터/정렬/검색/이미지 썸네일을 모두 렌더링한다.
    라이브러리 파일은 /static/tabulator/ 에서 서빙 (별도 복사 불필요).
    export-bundle ZIP에만 assets/로 포함된다.
    """
    html_dir = output_path.parent

    def e(v) -> str:
        return _html.escape("" if v is None else str(v))

    scenario_name = data.get("scenario_name", "")
    status = data.get("status", "")
    total_steps = data.get("total_steps", 0)
    total_repeat = data.get("total_repeat", 1)
    passed = data.get("passed_steps", 0)
    failed = data.get("failed_steps", 0)
    warned = data.get("warning_steps", 0)
    errored = data.get("error_steps", 0)
    started_at = data.get("started_at", "")
    finished_at = data.get("finished_at", "")

    def _fmt_ts(iso: str) -> str:
        try:
            from datetime import datetime as _dt
            ts = _dt.fromisoformat(iso.replace("Z", "+00:00"))
            return ts.astimezone().strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return iso or ""

    # Tabulator에 넘길 행 데이터 구성
    rows_json: list[dict] = []
    for sr in data.get("step_results", []):
        duration_ms = sr.get("execution_time_ms", 0) or 0
        delay_ms = sr.get("delay_ms", 0) or 0
        dur_str = f"{duration_ms}ms" if duration_ms < 1000 else f"{duration_ms / 1000:.1f}s"
        delay_str = f"{delay_ms}ms" if delay_ms else "-"
        exp_src = _html_image_src(sr.get("expected_image"), html_dir)
        act_src = _html_image_src(
            sr.get("actual_annotated_image") or sr.get("actual_image"), html_dir
        )
        rows_json.append({
            "timestamp": _fmt_ts(sr.get("timestamp", started_at)),
            "cycle": sr.get("repeat_index", 1),
            "step_id": sr.get("step_id", ""),
            "device": sr.get("device_id", ""),
            "command": sr.get("command", sr.get("message", "")),
            "description": sr.get("description", ""),
            "status": sr.get("status", ""),
            "delay": delay_str,
            "duration": dur_str,
            "expected_src": exp_src or "",
            "actual_src": act_src or "",
        })

    report_payload = {
        "scenario_name": scenario_name,
        "status": status,
        "total_repeat": total_repeat,
        "rows": rows_json,
    }
    # </script> 이스케이프 — 임베드 JSON 안에 script 종료 태그가 포함되면 파싱이 깨진다
    payload_json = json.dumps(report_payload, ensure_ascii=False).replace("</", "<\\/")

    parts: list[str] = []
    parts.append("<!DOCTYPE html>")
    parts.append('<html lang="ko"><head><meta charset="utf-8">')
    parts.append(f"<title>{e(scenario_name)} - Test Report</title>")
    parts.append('<script>var _tBase = location.protocol==="file:" ? "../../app/static/tabulator/" : "/static/tabulator/";</script>')
    parts.append('<script>document.write(\'<link rel="stylesheet" href="\'+_tBase+\'tabulator_simple.min.css">\');</script>')
    parts.append(f"<style>{_HTML_STYLE}</style>")
    parts.append("</head><body>")
    parts.append(f"<h1>{e(scenario_name)}</h1>")
    status_cls = status if status in ("pass", "fail", "error") else ""
    parts.append('<div class="meta">')
    parts.append(f'<span class="badge {status_cls}">{e(status.upper())}</span>')
    parts.append(f'<span class="stat">Step: {total_steps}</span>')
    parts.append(f'<span class="stat">Repeat: {total_repeat}</span>')
    parts.append(f'<span class="stat" style="background:#dcfce7;color:#15803d">Pass: {passed}</span>')
    if failed:
        parts.append(f'<span class="stat" style="background:#fee2e2;color:#b91c1c">Fail: {failed}</span>')
    if warned:
        parts.append(f'<span class="stat" style="background:#fef9c3;color:#854d0e">Warn: {warned}</span>')
    if errored:
        parts.append(f'<span class="stat" style="background:#ffedd5;color:#9a3412">Error: {errored}</span>')
    parts.append(f"<span>{e(_fmt_ts(started_at))} ~ {e(_fmt_ts(finished_at))}</span>")
    parts.append("</div>")

    # 상단 컨트롤 — 전역 검색, 필터 초기화, PDF 저장
    parts.append('<div class="controls">')
    parts.append('<input id="filter-text" type="text" placeholder="전체 검색 (모든 열)">')
    parts.append('<button id="filter-reset" class="btn-secondary" type="button">필터 초기화</button>')
    parts.append('<button id="pdf-btn" class="btn-primary" type="button">PDF 저장</button>')
    parts.append('<span class="count"><b id="filter-visible">0</b> / <b id="filter-total">0</b></span>')
    parts.append("</div>")

    # Tabulator 렌더 타겟
    parts.append('<div id="results-table"></div>')

    # 이미지 프리뷰 오버레이
    parts.append('<div id="preview-overlay" class="preview-overlay"><img id="preview-img" src="" alt=""></div>')

    # 데이터 임베드 + 라이브러리 로드 + 초기화
    parts.append(f'<script>window.__REPORT_DATA__ = {payload_json};</script>')
    parts.append('<script>document.write(\'<script src="\'+_tBase+\'tabulator.min.js"><\\/script>\');</script>')
    parts.append(f"<script>{_HTML_SCRIPT}</script>")
    parts.append("</body></html>")
    return "".join(parts)


def _build_excel_workbook(data: dict, filepath: Path = None):
    """Build an openpyxl Workbook from result data. Reusable by settings router."""
    import openpyxl
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    desc_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    desc_font = Font(color="44546A", size=9)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    vcenter = Alignment(vertical="center")
    vcenter_wrap = Alignment(vertical="center", wrap_text=True)

    col_headers = [
        "Time Stamp", "TOTAL TC REPEAT", "CURRENT TC REPEAT",
        "STEP INDEX", "Device", "Command", "Remark", "Status", "DELAY", "DURATION",
        "Expected Image", "Actual Image",
    ]
    col_descs = [
        "실행된 날짜/시간", "총 repeat", "현재 cycle",
        "스탭 순서", "장치", "action", "설명", "pass, fail, error, jump", "설정한 딜레이", "실제 걸린 시간",
        "기대 이미지", "비교 이미지 (annotated)",
    ]
    col_widths = [22, 16, 18, 12, 16, 30, 30, 12, 12, 14, 30, 30]

    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for ci, d in enumerate(col_descs, start=1):
        cell = ws.cell(row=2, column=ci, value=d)
        cell.font = desc_font
        cell.fill = desc_fill
        cell.alignment = center
        cell.border = thin_border

    total_repeat = data.get("total_repeat", 1)
    img_row_height = 120

    for ri, sr in enumerate(data.get("step_results", []), start=3):
        status = sr.get("status", "")
        timestamp = sr.get("timestamp", data.get("started_at", ""))
        command = sr.get("command", sr.get("message", ""))
        delay_ms = sr.get("delay_ms", 0)
        duration_ms = sr.get("execution_time_ms", 0)

        try:
            from datetime import datetime as _dt, timezone as _tz
            ts = _dt.fromisoformat(timestamp.replace("Z", "+00:00"))
            ts_local = ts.astimezone()  # 시스템 로컬 시간대로 변환
            ts_str = ts_local.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            ts_str = timestamp or ""

        dur_str = f"{duration_ms}ms" if duration_ms < 1000 else f"{duration_ms / 1000:.1f}s"
        delay_str = f"{delay_ms}ms" if delay_ms and delay_ms >= 1000 else (f"{delay_ms}ms" if delay_ms else "-")

        ws.cell(row=ri, column=1, value=ts_str).border = thin_border
        ws.cell(row=ri, column=1).alignment = center
        ws.cell(row=ri, column=2, value=total_repeat).border = thin_border
        ws.cell(row=ri, column=2).alignment = center
        ws.cell(row=ri, column=3, value=sr.get("repeat_index", 1)).border = thin_border
        ws.cell(row=ri, column=3).alignment = center
        ws.cell(row=ri, column=4, value=sr.get("step_id", "")).border = thin_border
        ws.cell(row=ri, column=4).alignment = center
        ws.cell(row=ri, column=5, value=sr.get("device_id", "")).border = thin_border
        ws.cell(row=ri, column=5).alignment = center
        ws.cell(row=ri, column=6, value=command).border = thin_border
        ws.cell(row=ri, column=6).alignment = vcenter_wrap
        ws.cell(row=ri, column=7, value=sr.get("description", "")).border = thin_border
        ws.cell(row=ri, column=7).alignment = vcenter_wrap
        status_cell = ws.cell(row=ri, column=8, value=status.upper())
        status_cell.border = thin_border
        status_cell.alignment = center
        if status == "pass":
            status_cell.fill = pass_fill
        elif status == "fail":
            status_cell.fill = fail_fill
        elif status == "warning":
            status_cell.fill = warn_fill
        elif status == "error":
            status_cell.fill = error_fill
        ws.cell(row=ri, column=9, value=delay_str).border = thin_border
        ws.cell(row=ri, column=9).alignment = center
        ws.cell(row=ri, column=10, value=dur_str).border = thin_border
        ws.cell(row=ri, column=10).alignment = center

        exp_path = _resolve_image_path(sr.get("expected_image"))
        ws.cell(row=ri, column=11).border = thin_border
        ws.cell(row=ri, column=11).alignment = center
        if exp_path:
            try:
                img = XlImage(str(exp_path))
                img.width = 180
                img.height = 140
                ws.add_image(img, f"K{ri}")
                ws.row_dimensions[ri].height = img_row_height
            except Exception:
                ws.cell(row=ri, column=11, value=str(sr.get("expected_image", "")))

        act_img_path = sr.get("actual_annotated_image") or sr.get("actual_image")
        act_path = _resolve_image_path(act_img_path)
        ws.cell(row=ri, column=12).border = thin_border
        ws.cell(row=ri, column=12).alignment = center
        if act_path:
            try:
                img = XlImage(str(act_path))
                img.width = 180
                img.height = 140
                ws.add_image(img, f"L{ri}")
                if ws.row_dimensions[ri].height is None or ws.row_dimensions[ri].height < img_row_height:
                    ws.row_dimensions[ri].height = img_row_height
            except Exception:
                ws.cell(row=ri, column=12, value=str(act_img_path or ""))

    return wb


@router.get("/export/{filename:path}")
async def export_result_excel(filename: str):
    """Export a test result as Excel (.xlsx) — download to browser."""
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    data = json.loads(filepath.read_text(encoding="utf-8"))

    try:
        wb = _build_excel_workbook(data, filepath)
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    export_name = filename.replace(".json", ".xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{export_name}"'},
    )


@router.post("/export-bundle/{filename:path}")
async def export_result_bundle(filename: str, export_path: str = ""):
    """결과 내보내기: 런 폴더를 ZIP으로 압축하여 다운로드 또는 지정 경로에 저장.

    - 런 폴더: 폴더 전체를 ZIP 압축
    - 레거시 파일: Excel + 녹화를 임시 폴더에 모아 ZIP 압축

    Args:
        export_path: 저장 경로. 빈 값이면 브라우저 다운로드.
    """
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    # 런 폴더인지 레거시인지 판별
    if filepath.name == "result.json" and filepath.parent != RESULTS_DIR:
        run_dir = filepath.parent
        folder_name = run_dir.name
        # 런 폴더에는 result.html만 자동 생성되고 xlsx는 없을 수 있다.
        # 번들 다운로드 시점에만 lazy로 xlsx를 생성하여 파일에 포함시킨다.
        excel_path = run_dir / "result.xlsx"
        if not excel_path.exists():
            try:
                data = json.loads(filepath.read_text(encoding="utf-8"))
                wb = _build_excel_workbook(data, filepath)
                wb.save(str(excel_path))
            except Exception:
                pass
    else:
        # 레거시: 임시 폴더에 결과물 수집
        data = json.loads(filepath.read_text(encoding="utf-8"))
        scenario_name = data.get("scenario_name", "unknown")
        started_at = data.get("started_at", "")
        try:
            dt = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
            ts = dt.strftime("%Y%m%d_%H%M%S")
        except Exception:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = scenario_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
        folder_name = f"{ts}_{safe_name}"

        import tempfile
        run_dir = Path(tempfile.mkdtemp()) / folder_name
        run_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(filepath), str(run_dir / filepath.name))

        # Excel 생성
        try:
            wb = _build_excel_workbook(data, filepath)
            wb.save(str(run_dir / filepath.name.replace(".json", ".xlsx")))
        except Exception:
            pass

        # 웹캠 녹화 복사 (webm + mp4)
        base = filename.replace(".json", "")
        if RECORDINGS_DIR.is_dir():
            for pattern in (f"{base}_webcam_*.webm", f"{base}_webcam_*.mp4"):
                for rec in sorted(RECORDINGS_DIR.glob(pattern)):
                    try:
                        shutil.copy2(str(rec), str(run_dir / rec.name))
                    except Exception:
                        pass

    # result.html이 /static/tabulator/ 절대경로를 참조하므로, ZIP 배포용으로
    # assets/를 런 폴더에 임시 복사 + HTML 내 경로를 상대경로로 패치한다.
    _tabulator_src = Path(__file__).resolve().parent.parent / "static" / "tabulator"
    _tmp_assets_dir = run_dir / "assets"
    _patched_html = False
    if _tabulator_src.is_dir():
        _tmp_assets_dir.mkdir(exist_ok=True)
        for _tf in ("tabulator.min.js", "tabulator_simple.min.css"):
            _src = _tabulator_src / _tf
            _dst = _tmp_assets_dir / _tf
            if _src.is_file() and not _dst.exists():
                shutil.copy2(str(_src), str(_dst))
        _html_file = run_dir / "result.html"
        if _html_file.is_file():
            _htxt = _html_file.read_text(encoding="utf-8")
            # 프로토콜 감지 로직을 ./assets/ 고정으로 교체
            _htxt_new = _htxt.replace(
                'var _tBase = location.protocol==="file:" ? "../../app/static/tabulator/" : "/static/tabulator/";',
                'var _tBase = "./assets/";',
            )
            if _htxt_new != _htxt:
                _html_file.write_text(_htxt_new, encoding="utf-8")
                _patched_html = True

    # ZIP 압축
    try:
        if export_path:
            zip_path = Path(export_path)
            if zip_path.is_dir():
                zip_path = zip_path / f"{folder_name}.zip"
            zip_path.parent.mkdir(parents=True, exist_ok=True)
            _zip_directory(run_dir, zip_path)
            return {"path": str(zip_path), "folder": folder_name, "size": zip_path.stat().st_size}
        else:
            buf = io.BytesIO()
            _zip_directory_to_buffer(run_dir, buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/zip",
                headers={"Content-Disposition": f'attachment; filename="{folder_name}.zip"'},
            )
    finally:
        # ZIP용 임시 assets 정리 + HTML 경로 복원 (런 폴더가 원본이면 패치 원복)
        if _patched_html:
            _html_file = run_dir / "result.html"
            if _html_file.is_file():
                _htxt = _html_file.read_text(encoding="utf-8")
                _html_file.write_text(
                    _htxt.replace(
                        'var _tBase = "./assets/";',
                        'var _tBase = location.protocol==="file:" ? "../../app/static/tabulator/" : "/static/tabulator/";',
                    ),
                    encoding="utf-8",
                )
        if _tmp_assets_dir.is_dir() and _tabulator_src.is_dir():
            shutil.rmtree(str(_tmp_assets_dir), ignore_errors=True)


@router.post("/open-folder")
async def open_result_folder(body: dict):
    """결과 폴더를 파일 탐색기로 열기."""
    import os, sys
    filename = body.get("filename", "")
    if not filename:
        raise HTTPException(status_code=400, detail="filename required")

    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    # 런 폴더면 그 폴더, 레거시면 RESULTS_DIR
    if filepath.name == "result.json" and filepath.parent != RESULTS_DIR:
        target = filepath.parent
    else:
        target = RESULTS_DIR

    if sys.platform == "win32":
        os.startfile(str(target))
    else:
        subprocess.Popen(["xdg-open", str(target)])
    return {"status": "ok", "path": str(target)}


def _iter_run_dir_files(source_dir: Path):
    """런 폴더 내 파일을 순회. junction/symlink 디렉토리는 실제 대상을 따라감."""
    for item in sorted(source_dir.rglob("*")):
        if item.is_file():
            yield item


def _zip_directory(source_dir: Path, zip_path: Path) -> None:
    """디렉토리를 ZIP 파일로 압축."""
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
        for file in _iter_run_dir_files(source_dir):
            arcname = file.relative_to(source_dir.parent).as_posix()
            zf.write(str(file), arcname)


def _zip_directory_to_buffer(source_dir: Path, buf: io.BytesIO) -> None:
    """디렉토리를 BytesIO 버퍼에 ZIP 압축."""
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in _iter_run_dir_files(source_dir):
            arcname = file.relative_to(source_dir.parent).as_posix()
            zf.write(str(file), arcname)


@router.delete("/{filename:path}")
async def delete_result(filename: str):
    """Delete a test result and its associated files.

    런 폴더(folder/result.json) 또는 레거시 플랫 파일(.json) 모두 처리.
    """
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    deleted_recordings = []

    # 런 폴더인 경우 폴더 전체 삭제
    if filepath.name == "result.json" and filepath.parent != RESULTS_DIR:
        run_dir = filepath.parent
        folder_name = run_dir.name
        shutil.rmtree(str(run_dir), ignore_errors=True)
        # 연결된 웹캠 녹화 파일도 삭제 (webm + mp4)
        if RECORDINGS_DIR.is_dir():
            for pattern in (f"{folder_name}_webcam_*.webm", f"{folder_name}_webcam_*.mp4"):
                for rec in RECORDINGS_DIR.glob(pattern):
                    rec.unlink()
                    deleted_recordings.append(rec.name)
    else:
        filepath.unlink()
        base = filename.replace(".json", "")
        if RECORDINGS_DIR.is_dir():
            for pattern in (f"{base}_webcam_*.webm", f"{base}_webcam_*.mp4"):
                for rec in RECORDINGS_DIR.glob(pattern):
                    rec.unlink()
                    deleted_recordings.append(rec.name)

    return {"status": "deleted", "deleted_recordings": deleted_recordings}


def _safe_filename(name: str) -> str:
    """Path traversal 방어: 파일명에서 디렉토리 부분 제거."""
    return Path(name).name


# --- Webcam recording endpoints ---

@router.post("/webcam-upload")
async def upload_webcam_recording(
    file: UploadFile = File(...),
    result_filename: str = Query(...),
    repeat_index: int = Query(1),
):
    """Upload a webcam recording linked to a test result."""
    base = result_filename.replace(".json", "").replace("/result", "")
    filename = f"webcam_r{repeat_index}.webm"
    content = await file.read()

    # 시나리오 결과 폴더의 recordings/ 에 저장
    run_dir = RESULTS_DIR / base
    if run_dir.is_dir():
        rec_dir = run_dir / "recordings"
        rec_dir.mkdir(exist_ok=True)
        filepath = rec_dir / filename
        filepath.write_bytes(content)
    else:
        # 결과 폴더가 없으면 기존 위치에 저장 (폴백)
        RECORDINGS_DIR.mkdir(parents=True, exist_ok=True)
        filepath = RECORDINGS_DIR / f"{base}_webcam_r{repeat_index}.webm"
        filepath.write_bytes(content)

    return {"filename": filename, "path": str(filepath)}


@router.get("/recordings-for/{result_filename:path}")
async def list_recordings_for_result(result_filename: str):
    """List webcam recordings linked to a test result (both .webm and .mp4)."""
    RECORDINGS_DIR.mkdir(parents=True, exist_ok=True)
    base = result_filename.replace(".json", "").replace("/result", "")
    recordings = []
    seen: set[str] = set()

    # 런 폴더 내 recordings/ 확인 (webm + mp4)
    run_dir = RESULTS_DIR / base
    rec_dir = run_dir / "recordings" if run_dir.is_dir() else None
    if rec_dir and rec_dir.is_dir():
        for pattern in ("*.webm", "*.mp4"):
            for f in sorted(rec_dir.glob(pattern)):
                if f.name in seen:
                    continue
                seen.add(f.name)
                recordings.append({
                    "filename": f.name,
                    "size": f.stat().st_size,
                    "url": f"/results-files/{base}/recordings/{f.name}",
                })

    # 레거시: Results/Video/ 에서도 탐색 (webm + mp4)
    for pattern in (f"{base}_webcam_*.webm", f"{base}_webcam_*.mp4"):
        for f in sorted(RECORDINGS_DIR.glob(pattern)):
            if f.name in seen:
                continue
            seen.add(f.name)
            recordings.append({
                "filename": f.name,
                "size": f.stat().st_size,
                "url": f"/recordings/{f.name}",
            })
    return {"recordings": recordings}


@router.delete("/recordings/{filename}")
async def delete_recording(filename: str):
    """Delete a webcam recording."""
    safe_name = _safe_filename(filename)
    filepath = RECORDINGS_DIR / safe_name
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Recording not found")
    filepath.unlink()
    return {"deleted": safe_name}


@router.post("/recordings/{filename}/trim")
async def trim_recording(
    filename: str,
    start: float = Query(...),
    end: float = Query(...),
):
    """Trim a webcam recording (requires ffmpeg)."""
    safe_name = _safe_filename(filename)
    filepath = RECORDINGS_DIR / safe_name
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Recording not found")
    if start >= end:
        raise HTTPException(status_code=400, detail="start must be less than end")
    ffmpeg_path = _find_ffmpeg()
    if ffmpeg_path is None:
        raise HTTPException(
            status_code=400,
            detail="ffmpeg가 설치되어 있지 않습니다. tools/ 폴더에 ffmpeg.exe를 넣거나 시스템에 설치하세요."
        )
    output_name = f"trim_{start:.1f}_{end:.1f}_{safe_name}"
    output_path = RECORDINGS_DIR / output_name
    try:
        subprocess.run(
            [ffmpeg_path, "-i", str(filepath), "-ss", str(start), "-to", str(end),
             "-c", "copy", str(output_path), "-y"],
            check=True, capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
        )
    except subprocess.CalledProcessError as e:
        raise HTTPException(status_code=500, detail=f"ffmpeg error: {e.stderr.decode(errors='replace')[:300]}")
    return {"filename": output_name, "url": f"/recordings/{output_name}"}


@router.post("/update-step/{filename:path}")
async def update_step_result(filename: str, body: dict):
    """백그라운드 CMD 완료 후 스텝 결과를 영구 업데이트."""
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")
    data = json.loads(filepath.read_text(encoding="utf-8"))
    step_index = body.get("step_index")
    if step_index is None or step_index < 0 or step_index >= len(data.get("step_results", [])):
        raise HTTPException(status_code=400, detail="Invalid step_index")

    sr = data["step_results"][step_index]
    if "message" in body:
        sr["message"] = body["message"]
    if "status" in body:
        old_status = sr["status"]
        new_status = body["status"]
        sr["status"] = new_status
        # 카운트 재계산
        if old_status != new_status:
            status_map = {"pass": "passed_steps", "fail": "failed_steps",
                          "warning": "warning_steps", "error": "error_steps"}
            if old_status in status_map:
                data[status_map[old_status]] = max(0, data.get(status_map[old_status], 0) - 1)
            if new_status in status_map:
                data[status_map[new_status]] = data.get(status_map[new_status], 0) + 1
            # 전체 상태 재평가
            if data.get("failed_steps", 0) > 0 or data.get("error_steps", 0) > 0:
                data["status"] = "fail"
            elif data.get("warning_steps", 0) > 0:
                data["status"] = "warning"
            else:
                data["status"] = "pass"

    filepath.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"status": "ok", "result_status": data["status"]}


@router.post("/migrate-legacy")
async def migrate_legacy():
    """레거시 결과 파일을 새 구조로 마이그레이션.
    screenshots/{name}/actual_{ts}/ → results/{ts}_{name}/screenshots/
    results/{name}_{ts}.json → results/{ts}_{name}/result.json
    """
    import re as _re
    migrated = 0
    errors = []

    # 1) screenshots 내 actual_ 폴더 → results 런 폴더로 이동
    if SCREENSHOTS_DIR.is_dir():
        for scenario_dir in SCREENSHOTS_DIR.iterdir():
            if not scenario_dir.is_dir():
                continue
            sc_name = scenario_dir.name
            for actual_dir in list(scenario_dir.iterdir()):
                if not actual_dir.is_dir() or not actual_dir.name.startswith("actual_"):
                    continue
                ts = actual_dir.name.replace("actual_", "")  # e.g. 20260408_174101
                if not _re.match(r"\d{8}_\d{6}", ts):
                    continue
                safe_name = _re.sub(r'[\\/:*?"<>|→]', '_', sc_name).replace(" ", "_")
                run_dir = RESULTS_DIR / f"{ts}_{safe_name}"
                run_dir.mkdir(parents=True, exist_ok=True)
                dst_ss = run_dir / "screenshots"
                if not dst_ss.exists():
                    try:
                        shutil.move(str(actual_dir), str(dst_ss))
                        migrated += 1
                    except Exception as e:
                        errors.append(f"screenshots/{sc_name}/{actual_dir.name}: {e}")
                else:
                    # 이미 존재하면 파일 단위로 머지
                    for f in actual_dir.iterdir():
                        if f.is_file():
                            dst_f = dst_ss / f.name
                            if not dst_f.exists():
                                shutil.move(str(f), str(dst_f))
                    # 빈 폴더 삭제
                    try:
                        actual_dir.rmdir()
                    except Exception:
                        pass
                    migrated += 1

    # 2) results 내 플랫 JSON → 런 폴더로 이동
    if RESULTS_DIR.is_dir():
        for json_file in list(RESULTS_DIR.glob("*.json")):
            # {name}_{timestamp}.json 패턴 매칭
            m = _re.match(r"^(.+?)_(\d{8}_\d{6})\.json$", json_file.name)
            if not m:
                continue
            sc_name = m.group(1)
            ts = m.group(2)
            safe_name = _re.sub(r'[\\/:*?"<>|→]', '_', sc_name).replace(" ", "_")
            run_dir = RESULTS_DIR / f"{ts}_{safe_name}"
            run_dir.mkdir(parents=True, exist_ok=True)
            dst = run_dir / "result.json"
            if not dst.exists():
                try:
                    shutil.move(str(json_file), str(dst))
                    # Excel도 함께 이동
                    xlsx = json_file.with_suffix(".xlsx")
                    if xlsx.exists():
                        shutil.move(str(xlsx), str(run_dir / "result.xlsx"))
                    migrated += 1
                except Exception as e:
                    errors.append(f"{json_file.name}: {e}")

    # 3) screenshots 내 actual/actual_ 폴더 정리 + 빈 폴더 삭제
    if SCREENSHOTS_DIR.is_dir():
        for d in list(SCREENSHOTS_DIR.iterdir()):
            if not d.is_dir():
                continue
            for sub in list(d.iterdir()):
                if sub.is_dir() and sub.name == "actual":
                    # 타임스탬프 없는 actual 폴더 (단일 스텝 테스트 임시) → 삭제
                    try:
                        shutil.rmtree(str(sub))
                        migrated += 1
                    except Exception as e:
                        errors.append(f"screenshots/{d.name}/actual: {e}")
            # 하위에 actual_ 폴더도 파일도 없으면 폴더 자체 삭제
            try:
                remaining = list(d.iterdir())
                if not remaining:
                    d.rmdir()
            except Exception:
                pass

    return {"migrated": migrated, "errors": errors}


@router.get("/{filename:path}")
async def get_result(filename: str):
    """Get a specific test result (런 폴더 또는 레거시 플랫 파일)."""
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")
    data = json.loads(filepath.read_text(encoding="utf-8"))
    return data


@router.get("/image/{scenario_name}/{image_path:path}")
async def get_image(scenario_name: str, image_path: str):
    """Serve a screenshot image."""
    filepath = SCREENSHOTS_DIR / scenario_name / image_path
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(str(filepath), media_type="image/png")
